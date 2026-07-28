from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from aglae.traupixe.format import TRAUPIXE_FORMAT
from aglae.traupixe.models import Detector, MeasurementUnit

DEFAULT_ANALYSIS_IDS = ("opaque-analysis-a", "opaque-analysis-b")

MeasurementKey = tuple[str, str, MeasurementUnit]
DetectorKey = tuple[str, str]
HeaderKey = tuple[str, int]


def write_traupixe_fixture(
    path: Path,
    *,
    analysis_ids: Sequence[object] = DEFAULT_ANALYSIS_IDS,
    row_orders: Mapping[str, Sequence[object]] | None = None,
    values: Mapping[MeasurementKey, tuple[object, object]] | None = None,
    detectors: Mapping[DetectorKey, object] | None = None,
    header_overrides: Mapping[HeaderKey, object] | None = None,
    missing_sheet: str | None = None,
    extra_sheet: str | None = None,
    force_a1_dimensions: bool = False,
) -> Path:
    row_orders = row_orders or {}
    values = values or {}
    detectors = detectors or {}
    header_overrides = header_overrides or {}

    workbook = Workbook()
    workbook.remove(workbook.active)
    for worksheet_format in TRAUPIXE_FORMAT.worksheets:
        if worksheet_format.name == missing_sheet:
            continue
        worksheet = workbook.create_sheet(worksheet_format.name)
        if worksheet_format.header_row is not None:
            for column, header in enumerate(
                worksheet_format.headers,
                start=1,
            ):
                worksheet.cell(
                    row=worksheet_format.header_row,
                    column=column,
                    value=header_overrides.get(
                        (worksheet_format.name, column),
                        header,
                    ),
                )

    if extra_sheet is not None:
        workbook.create_sheet(extra_sheet)

    descriptions = {
        analysis_id: f"Description {index}"
        for index, analysis_id in enumerate(analysis_ids, start=1)
    }
    for sheet_name in TRAUPIXE_FORMAT.source_sheets:
        worksheet = workbook[sheet_name]
        worksheet_format = TRAUPIXE_FORMAT.worksheet(sheet_name)
        assert worksheet_format.data_start_row is not None
        sheet_analysis_ids = row_orders.get(sheet_name, analysis_ids)
        for offset, analysis_id in enumerate(sheet_analysis_ids):
            row = worksheet_format.data_start_row + offset
            worksheet.cell(row=row, column=1, value=analysis_id)
            worksheet.cell(
                row=row,
                column=2,
                value=descriptions.get(
                    analysis_id,
                    f"Description {analysis_id}",
                ),
            )

            if sheet_name == "S_Best Det.":
                for analyte_index, analyte in enumerate(
                    TRAUPIXE_FORMAT.analytes,
                ):
                    worksheet.cell(
                        row=row,
                        column=analyte_index + 3,
                        value=detectors.get(
                            (str(analysis_id), analyte),
                            (
                                Detector.X0.value
                                if analyte_index % 2 == 0
                                else Detector.X10.value
                            ),
                        ),
                    )
                continue

            unit = dict(TRAUPIXE_FORMAT.unit_sheets).get(sheet_name)
            if unit is None:
                continue
            for analyte_index, analyte in enumerate(
                TRAUPIXE_FORMAT.analytes,
            ):
                default_value = "1,25" if unit is MeasurementUnit.PERCENT else "12500"
                raw_value, raw_uncertainty = values.get(
                    (str(analysis_id), analyte, unit),
                    (default_value, "2,5"),
                )
                value_column = 3 + (analyte_index * 2)
                worksheet.cell(
                    row=row,
                    column=value_column,
                    value=raw_value,
                )
                worksheet.cell(
                    row=row,
                    column=value_column + 1,
                    value=raw_uncertainty,
                )

    workbook.save(path)
    if force_a1_dimensions:
        _replace_dimensions_with_a1(path)
    return path


def _replace_dimensions_with_a1(path: Path) -> None:
    rewritten_path = path.with_suffix(".rewritten.xlsx")
    with ZipFile(path, "r") as source_archive, ZipFile(
        rewritten_path,
        "w",
        compression=ZIP_DEFLATED,
    ) as target_archive:
        for item in source_archive.infolist():
            data = source_archive.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                data = re.sub(
                    rb'<dimension ref="[^"]+"\s*/>',
                    b'<dimension ref="A1"/>',
                    data,
                    count=1,
                )
            target_archive.writestr(item, data)
    rewritten_path.replace(path)
