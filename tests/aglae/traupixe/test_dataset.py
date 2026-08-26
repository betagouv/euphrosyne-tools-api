from io import BytesIO

import pytest
from openpyxl import Workbook

from aglae.traupixe.dataset import (
    TraupixeDatasetError,
    load_traupixe_dataset,
    serialize_traupixe_for_model,
)


def test_interprets_concentrations_detections_groups_and_detectors(
    traupixe_workbook: bytes,
) -> None:
    dataset = load_traupixe_dataset(traupixe_workbook)

    assert dataset.concentration_sheet == "S_Conc. %"
    assert dataset.detector_sheet == "S_Best Det."
    assert dataset.unit == "%"
    assert len(dataset.analyses) == 5
    assert len(dataset.non_reference_analyses) == 4
    assert [analysis.group for analysis in dataset.non_reference_analyses] == [
        "Zone_A",
        "Zone_A",
        "Zone_B",
        "Zone_B",
    ]
    assert dataset.analyses[0].kind == "reference"
    assert all(
        analysis.kind == "unknown" for analysis in dataset.non_reference_analyses
    )
    zone_a_second = dataset.non_reference_analyses[1]
    assert zone_a_second.measurements["Na2O-PIGE"].value == 4
    assert zone_a_second.measurements["Zr"].detected is False
    assert zone_a_second.measurements["Zr"].value == pytest.approx(0.2)
    assert zone_a_second.measurements["Zr"].detector in {"X1", "X3"}


def test_builds_a_compact_descriptor_without_measurement_values(
    traupixe_workbook: bytes,
) -> None:
    descriptor = load_traupixe_dataset(traupixe_workbook).descriptor()

    assert descriptor["analyses"]["references"] == 1
    assert descriptor["analyses"]["unknown"] == 4
    assert "default_major_analytes" not in descriptor
    assert descriptor["analyses"]["groups"] == [
        {"name": "Zone_A", "analyses": 2},
        {"name": "Zone_B", "analyses": 2},
    ]
    by_analyte = {analyte["name"]: analyte for analyte in descriptor["analytes"]}
    assert by_analyte["Zr"]["detection_rate"] == 0.75
    assert by_analyte["Ge"]["detection_rate"] == 0.25
    assert "value" not in str(descriptor)


def test_serializes_an_aligned_compact_model_payload(
    traupixe_workbook: bytes,
) -> None:
    dataset = load_traupixe_dataset(traupixe_workbook)

    payload = serialize_traupixe_for_model(dataset)

    assert payload["unit"] == "%"
    assert payload["summary"]["unknown"] == 4
    assert "default_major_analytes" not in payload
    assert len(payload["analyses"]) == len(dataset.analyses)
    assert all(
        len(analysis[key]) == len(payload["analytes"])
        for analysis in payload["analyses"]
        for key in ("values", "detected", "detectors")
    )
    assert {analysis["kind"] for analysis in payload["analyses"]} == {
        "reference",
        "unknown",
    }


def test_header_must_leave_identifier_and_label_columns_empty() -> None:
    workbook = Workbook()
    concentrations = workbook.active
    assert concentrations is not None
    concentrations.title = "S_Conc. %"
    concentrations.append([None, None, "A", "B", "C"])
    concentrations.append(["same-id", "Zone pt1", "1", "2", "3", "x", "y"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    dataset = load_traupixe_dataset(output.getvalue())

    assert dataset.analytes == ("A", "B", "C")
    assert dataset.analyses[0].group == "Zone"


def test_matches_detectors_by_identifier_occurrence_with_labeled_header() -> None:
    workbook = Workbook()
    concentrations = workbook.active
    assert concentrations is not None
    concentrations.title = "S_Conc. %"
    concentrations.append([None, None, "A", "B", "C"])
    concentrations.append(["same-id", "Zone_pt1", "1", "2", "3"])
    concentrations.append(["same-id", "Zone pt2", "4", "5", "6"])
    detectors = workbook.create_sheet("S_Best Det.")
    detectors.append(["Run", "Infos run", "A", "B", "C"])
    detectors.append(["same-id", "First label", "X0", "X0", "X0"])
    detectors.append(["same-id", "Truncated label", "X3", "X3", "X3"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    dataset = load_traupixe_dataset(output.getvalue())

    assert [analysis.group for analysis in dataset.analyses] == ["Zone", "Zone"]
    assert [analysis.measurements["A"].detector for analysis in dataset.analyses] == [
        "X0",
        "X3",
    ]


def test_rejects_a_workbook_without_a_supported_concentration_sheet() -> None:
    workbook = Workbook()
    active_sheet = workbook.active
    assert active_sheet is not None
    active_sheet.title = "Other"
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(TraupixeDatasetError, match="concentration sheet"):
        load_traupixe_dataset(output.getvalue())
