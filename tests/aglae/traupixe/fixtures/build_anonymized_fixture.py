from __future__ import annotations

import argparse
import hashlib
import re
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

REFERENCE_SHA256 = "bf7861fb9bc2d4ee43951fffa02281b01c2676c8c04d5fecdd246b27ae1a56b0"
SOURCE_SHEET = "S_Conc. & Unc %"
EXPERIMENT_SHEET = "Exp. data"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _rewrite_dimensions_as_a1(source: Path, destination: Path) -> None:
    dimension_pattern = re.compile(rb'<dimension ref="[^"]+"\s*/>')
    with (
        zipfile.ZipFile(source, "r") as input_archive,
        zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as output_archive,
    ):
        for entry in input_archive.infolist():
            content = input_archive.read(entry.filename)
            if entry.filename.startswith(
                "xl/worksheets/sheet"
            ) and entry.filename.endswith(".xml"):
                content = dimension_pattern.sub(b'<dimension ref="A1:A1"/>', content, 1)
            output_archive.writestr(entry, content)


def anonymize_fixture(source: Path, destination: Path) -> None:
    if _sha256(source) != REFERENCE_SHA256:
        raise ValueError("The input does not match the documented reference workbook.")

    workbook = load_workbook(source, read_only=False, data_only=False, keep_links=False)
    source_sheet = workbook[SOURCE_SHEET]
    analysis_ids = [
        str(source_sheet.cell(row=row, column=1).value)
        for row in range(3, source_sheet.max_row + 1)
        if source_sheet.cell(row=row, column=1).value
    ]
    if len(analysis_ids) != 48 or len(set(analysis_ids)) != 48:
        raise ValueError("The reference workbook must contain 48 unique analyses.")

    identifiers = {
        analysis_id: f"ANALYSIS-{index:04d}"
        for index, analysis_id in enumerate(analysis_ids, start=1)
    }
    descriptions = {
        analysis_id: f"Analyse anonymisée {index:04d}"
        for index, analysis_id in enumerate(analysis_ids, start=1)
    }

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            row_analysis_id = row[0].value if row else None
            if isinstance(row_analysis_id, str) and row_analysis_id in identifiers:
                original_id = row_analysis_id
                row[0].value = identifiers[original_id]
                if len(row) > 1:
                    row[1].value = descriptions[original_id]
                if worksheet.title == EXPERIMENT_SHEET:
                    for cell in row[2:]:
                        cell.value = None
                continue

            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                anonymized = cell.value
                for original_id, replacement in identifiers.items():
                    anonymized = anonymized.replace(original_id, replacement)
                cell.value = anonymized

    workbook.properties.creator = "Euphrosyne"
    workbook.properties.lastModifiedBy = "Euphrosyne"
    workbook.properties.title = "TRAUPIXE anonymized test fixture"
    workbook.properties.subject = None
    workbook.properties.description = (
        "Anonymized fixture derived from the documented TRAUPIXE reference workbook."
    )
    workbook.properties.keywords = None
    workbook.properties.category = "Test fixture"

    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="traupixe-fixture-") as temporary:
        normalized = Path(temporary) / "normalized.xlsx"
        workbook.save(normalized)
        _rewrite_dimensions_as_a1(normalized, destination)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    anonymize_fixture(args.source, args.destination)


if __name__ == "__main__":
    main()
