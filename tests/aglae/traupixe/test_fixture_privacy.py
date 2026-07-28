from pathlib import Path

from openpyxl import load_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "traupixe_reference_anonymized.xlsx"


def test_reference_fixture_is_anonymized_and_keeps_48_analyses():
    workbook = load_workbook(FIXTURE, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == "Informations":
                continue
            worksheet.reset_dimensions()
            worksheet.calculate_dimension(force=True)

        source = workbook["S_Conc. & Unc %"]
        rows = [
            row
            for row in source.iter_rows(min_row=3, values_only=True)
            if row[0] is not None
        ]

        assert len(rows) == 48
        assert [row[0] for row in rows] == [
            f"ANALYSIS-{index:04d}" for index in range(1, 49)
        ]
        assert [row[1] for row in rows] == [
            f"Analyse anonymisée {index:04d}" for index in range(1, 49)
        ]

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str):
                        assert "20241203_" not in value
                        assert "_TACT_IBA" not in value
    finally:
        workbook.close()
