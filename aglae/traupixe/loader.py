from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
from typing import Any, BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .exceptions import (
    TraupixeIncompatibleWorkbookError,
    TraupixeTooLargeError,
    TraupixeUnreadableError,
    TraupixeUnsupportedFileError,
    TraupixeValidationCode,
    TraupixeValidationIssue,
)
from .format import TRAUPIXE_FORMAT, TraupixeFormat, WorksheetFormat
from .models import Analysis, Detector, MeasurementUnit

WorkbookSource = str | Path | BinaryIO
AnalysisKey = tuple[str, int]
IdentifiedRow = tuple[int, AnalysisKey, tuple[object, ...]]


@dataclass(frozen=True)
class LoadedMeasurement:
    analysis_id: str
    analyte: str
    unit: MeasurementUnit
    raw_value: object
    raw_uncertainty: object
    detector: Detector | None


@dataclass(frozen=True)
class LoadedTraupixeWorkbook:
    source_name: str
    source_sha256: str
    analyses: tuple[Analysis, ...]
    measurements: tuple[LoadedMeasurement, ...]
    analytes: tuple[str, ...]
    units: tuple[MeasurementUnit, ...]
    detectors: tuple[Detector, ...]


@dataclass(frozen=True)
class _SourceDetails:
    name: str
    size: int
    sha256: str
    initial_position: int | None = None


def _read_source_details(
    source: WorkbookSource,
    source_name: str | None,
    maximum_size: int,
) -> _SourceDetails:
    if isinstance(source, (str, Path)):
        path = Path(source)
        name = source_name or path.name
        reported_size = path.stat().st_size
        if reported_size > maximum_size:
            raise TraupixeTooLargeError(reported_size, maximum_size)
        with path.open("rb") as source_file:
            size, source_sha256 = _bounded_source_fingerprint(
                source_file,
                maximum_size,
            )
        return _SourceDetails(
            name=name,
            size=size,
            sha256=source_sha256,
        )

    if not source.seekable():
        raise TraupixeUnreadableError(
            "A seekable stream is required to read a TRAUPIXE workbook"
        )

    initial_position = source.tell()
    source.seek(0)
    try:
        size, source_sha256 = _bounded_source_fingerprint(
            source,
            maximum_size,
        )
    except Exception:
        source.seek(initial_position)
        raise
    else:
        source.seek(0)

    stream_name = source_name
    if stream_name is None:
        raw_name = getattr(source, "name", "source.xlsx")
        stream_name = Path(str(raw_name)).name

    return _SourceDetails(
        name=stream_name,
        size=size,
        sha256=source_sha256,
        initial_position=initial_position,
    )


def _bounded_source_fingerprint(
    source: BinaryIO,
    maximum_size: int,
) -> tuple[int, str]:
    digest = sha256()
    size = 0
    while True:
        read_size = min(1024 * 1024, maximum_size - size + 1)
        chunk = source.read(read_size)
        if not chunk:
            return size, digest.hexdigest()
        size += len(chunk)
        if size > maximum_size:
            raise TraupixeTooLargeError(size, maximum_size)
        digest.update(chunk)


def _restore_stream_position(
    source: WorkbookSource,
    initial_position: int | None,
) -> None:
    if isinstance(source, (str, Path)) or initial_position is None:
        return
    source.seek(initial_position)


def _prepare_worksheet(worksheet: Any) -> None:
    dimension = worksheet.calculate_dimension()
    if dimension not in {"A1", "A1:A1"}:
        return
    worksheet.reset_dimensions()
    try:
        worksheet.calculate_dimension(force=True)
    except UnboundLocalError:
        # openpyxl raises this for a genuinely empty read-only worksheet.
        return


def _validate_sheet_names(
    sheet_names: tuple[str, ...],
    traupixe_format: TraupixeFormat,
) -> list[TraupixeValidationIssue]:
    actual = set(sheet_names)
    return [
        TraupixeValidationIssue(
            code=TraupixeValidationCode.MISSING_SHEET,
            message=f"Missing required worksheet: {sheet_name}",
            sheet=sheet_name,
        )
        for sheet_name in traupixe_format.required_sheets
        if sheet_name not in actual
    ]


def _trim_trailing_empty(values: tuple[object, ...]) -> tuple[object, ...]:
    end = len(values)
    while end and values[end - 1] is None:
        end -= 1
    return values[:end]


def _read_header_values(
    worksheet: Any,
    worksheet_format: WorksheetFormat,
) -> tuple[object, ...]:
    row = next(
        worksheet.iter_rows(
            min_row=worksheet_format.header_row,
            max_row=worksheet_format.header_row,
            max_col=worksheet.max_column or 1,
            values_only=True,
        ),
        (),
    )
    return _trim_trailing_empty(tuple(row))


def _selected_analytes(
    worksheet: Any,
    worksheet_format: WorksheetFormat,
    traupixe_format: TraupixeFormat,
) -> tuple[tuple[str, ...], list[TraupixeValidationIssue]]:
    header_values = _read_header_values(worksheet, worksheet_format)
    selected_values = header_values[2:]
    issues: list[TraupixeValidationIssue] = []
    if not selected_values or len(selected_values) % 2:
        return (), [
            TraupixeValidationIssue(
                code=TraupixeValidationCode.INVALID_HEADER,
                message=(
                    "Expected one or more analyte/Unc% column pairs "
                    "starting in column C"
                ),
                sheet=worksheet.title,
                cell=f"C{worksheet_format.header_row}",
            )
        ]

    analytes: list[str] = []
    for offset in range(0, len(selected_values), 2):
        raw_analyte = selected_values[offset]
        raw_uncertainty = selected_values[offset + 1]
        analyte_column = offset + 3
        uncertainty_column = analyte_column + 1
        if not isinstance(raw_analyte, str) or not raw_analyte.strip():
            issues.append(
                TraupixeValidationIssue(
                    code=TraupixeValidationCode.INVALID_HEADER,
                    message="Analyte headers must be non-empty text",
                    sheet=worksheet.title,
                    cell=(
                        f"{_column_letter(analyte_column)}"
                        f"{worksheet_format.header_row}"
                    ),
                )
            )
        else:
            analytes.append(traupixe_format.canonicalize_header(raw_analyte))
        if (
            not isinstance(raw_uncertainty, str)
            or raw_uncertainty.strip().casefold() != "unc%"
        ):
            issues.append(
                TraupixeValidationIssue(
                    code=TraupixeValidationCode.INVALID_HEADER,
                    message="Each analyte column must be followed by Unc%",
                    sheet=worksheet.title,
                    cell=(
                        f"{_column_letter(uncertainty_column)}"
                        f"{worksheet_format.header_row}"
                    ),
                )
            )

    duplicates = sorted(
        analyte for analyte in set(analytes) if analytes.count(analyte) > 1
    )
    if duplicates:
        issues.append(
            TraupixeValidationIssue(
                code=TraupixeValidationCode.INVALID_HEADER,
                message="Duplicate analyte headers: " + ", ".join(duplicates),
                sheet=worksheet.title,
            )
        )
    return tuple(analytes), issues


def _detector_analytes(
    worksheet: Any,
    worksheet_format: WorksheetFormat,
    traupixe_format: TraupixeFormat,
) -> tuple[tuple[str, ...], list[TraupixeValidationIssue]]:
    header_values = _read_header_values(worksheet, worksheet_format)[2:]
    analytes: list[str] = []
    issues: list[TraupixeValidationIssue] = []
    if not header_values:
        issues.append(
            TraupixeValidationIssue(
                code=TraupixeValidationCode.INVALID_HEADER,
                message="Expected one or more analyte columns starting in column C",
                sheet=worksheet.title,
                cell=f"C{worksheet_format.header_row}",
            )
        )
    for offset, raw_analyte in enumerate(header_values, start=3):
        if not isinstance(raw_analyte, str) or not raw_analyte.strip():
            issues.append(
                TraupixeValidationIssue(
                    code=TraupixeValidationCode.INVALID_HEADER,
                    message="Analyte headers must be non-empty text",
                    sheet=worksheet.title,
                    cell=f"{_column_letter(offset)}{worksheet_format.header_row}",
                )
            )
            continue
        analytes.append(traupixe_format.canonicalize_header(raw_analyte))
    return tuple(analytes), issues


def _discover_analytes(
    workbook: Any,
    traupixe_format: TraupixeFormat,
) -> tuple[tuple[str, ...], list[TraupixeValidationIssue]]:
    sequences: dict[str, tuple[str, ...]] = {}
    issues: list[TraupixeValidationIssue] = []
    for sheet_name, _ in traupixe_format.unit_sheets:
        worksheet_format = traupixe_format.worksheet(sheet_name)
        sequence, sequence_issues = _selected_analytes(
            workbook[sheet_name],
            worksheet_format,
            traupixe_format,
        )
        sequences[sheet_name] = sequence
        issues.extend(sequence_issues)

    detector_sheet = "S_Best Det."
    detector_sequence, detector_issues = _detector_analytes(
        workbook[detector_sheet],
        traupixe_format.worksheet(detector_sheet),
        traupixe_format,
    )
    sequences[detector_sheet] = detector_sequence
    issues.extend(detector_issues)

    reference_sheet = traupixe_format.unit_sheets[0][0]
    reference = sequences[reference_sheet]
    for sheet_name, sequence in sequences.items():
        if sheet_name == reference_sheet or sequence == reference:
            continue
        issues.append(
            TraupixeValidationIssue(
                code=TraupixeValidationCode.INVALID_HEADER,
                message=(f"Analyte sequence does not match {reference_sheet}"),
                sheet=sheet_name,
            )
        )
    return reference, issues


def _identified_rows(
    worksheet: Any,
    worksheet_format: WorksheetFormat,
    *,
    maximum_column: int,
) -> tuple[
    tuple[IdentifiedRow, ...],
    list[TraupixeValidationIssue],
]:
    rows: list[IdentifiedRow] = []
    issues: list[TraupixeValidationIssue] = []
    occurrences: dict[str, int] = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=worksheet_format.data_start_row,
            max_col=maximum_column,
            values_only=True,
        ),
        start=worksheet_format.data_start_row,
    ):
        analysis_id = row[0] if row else None
        if analysis_id is None or (
            isinstance(analysis_id, str) and not analysis_id.strip()
        ):
            continue
        if not isinstance(analysis_id, str):
            issues.append(
                TraupixeValidationIssue(
                    code=TraupixeValidationCode.INVALID_ANALYSIS_ID,
                    message="Analysis identifiers must be non-empty text",
                    sheet=worksheet.title,
                    cell=f"A{row_number}",
                )
            )
            continue
        description = row[1] if len(row) > 1 else None
        if not isinstance(description, str):
            issues.append(
                TraupixeValidationIssue(
                    code=TraupixeValidationCode.INVALID_VALUE,
                    message="Analysis descriptions must be text",
                    sheet=worksheet.title,
                    cell=f"B{row_number}",
                )
            )
        occurrence = occurrences.get(analysis_id, 0) + 1
        occurrences[analysis_id] = occurrence
        rows.append((row_number, (analysis_id, occurrence), tuple(row)))
    return tuple(rows), issues


def _analysis_key_text(analysis_key: AnalysisKey) -> str:
    analysis_id, occurrence = analysis_key
    if occurrence == 1:
        return analysis_id
    return f"{analysis_id} (occurrence {occurrence})"


def _validate_aligned_ids(
    rows_by_sheet: dict[
        str,
        tuple[IdentifiedRow, ...],
    ],
    source_sheets: tuple[str, ...],
) -> list[TraupixeValidationIssue]:
    reference_sheet = source_sheets[0]
    reference_rows = {
        analysis_key: (row_number, row)
        for row_number, analysis_key, row in rows_by_sheet[reference_sheet]
    }
    reference_ids = set(reference_rows)
    issues: list[TraupixeValidationIssue] = []

    if not reference_ids:
        issues.append(
            TraupixeValidationIssue(
                code=TraupixeValidationCode.INVALID_ANALYSIS_ID,
                message="The TRAUPIXE workbook contains no analyses",
                sheet=reference_sheet,
            )
        )

    for sheet_name in source_sheets[1:]:
        actual_rows = {
            analysis_key: (row_number, row)
            for row_number, analysis_key, row in rows_by_sheet[sheet_name]
        }
        actual_ids = set(actual_rows)
        missing = sorted(reference_ids - actual_ids)
        unexpected = sorted(actual_ids - reference_ids)
        if missing or unexpected:
            details = []
            if missing:
                details.append(
                    "missing: " + ", ".join(_analysis_key_text(key) for key in missing)
                )
            if unexpected:
                details.append(
                    "unexpected: "
                    + ", ".join(_analysis_key_text(key) for key in unexpected)
                )
            issues.append(
                TraupixeValidationIssue(
                    code=TraupixeValidationCode.MISALIGNED_ANALYSIS_IDS,
                    message=(
                        f"Analysis identifiers do not match {reference_sheet}"
                        + (f" ({'; '.join(details)})" if details else "")
                    ),
                    sheet=sheet_name,
                )
            )
            continue

        for analysis_key in reference_ids:
            expected_description = reference_rows[analysis_key][1][1]
            row_number, actual_row = actual_rows[analysis_key]
            if actual_row[1] != expected_description:
                issues.append(
                    TraupixeValidationIssue(
                        code=TraupixeValidationCode.INVALID_VALUE,
                        message=(
                            "Analysis description does not match "
                            f"{reference_sheet} for "
                            f"{_analysis_key_text(analysis_key)}"
                        ),
                        sheet=sheet_name,
                        cell=f"B{row_number}",
                    )
                )
    return issues


def _validate_measurement_value_types(
    rows_by_sheet: dict[
        str,
        tuple[IdentifiedRow, ...],
    ],
    traupixe_format: TraupixeFormat,
) -> list[TraupixeValidationIssue]:
    issues: list[TraupixeValidationIssue] = []
    for sheet_name, _ in traupixe_format.unit_sheets:
        for row_number, _, row in rows_by_sheet[sheet_name]:
            for column_number, value in enumerate(row[2:], start=3):
                if value is None or isinstance(value, str):
                    continue
                issues.append(
                    TraupixeValidationIssue(
                        code=TraupixeValidationCode.INVALID_VALUE,
                        message=(
                            "Selected concentrations and uncertainties "
                            "must be text or empty"
                        ),
                        sheet=sheet_name,
                        cell=f"{_column_letter(column_number)}{row_number}",
                    )
                )
    return issues


def _validate_no_formulas(
    workbook: Any,
    traupixe_format: TraupixeFormat,
) -> list[TraupixeValidationIssue]:
    issues: list[TraupixeValidationIssue] = []
    for worksheet_format in traupixe_format.worksheets:
        worksheet = workbook[worksheet_format.name]
        _prepare_worksheet(worksheet)
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                issues.append(
                    TraupixeValidationIssue(
                        code=TraupixeValidationCode.INVALID_VALUE,
                        message="Formulas are not allowed in TRAUPIXE workbooks",
                        sheet=worksheet.title,
                        cell=cell.coordinate,
                    )
                )
    return issues


def _normalize_detector(
    raw_detector: object,
    *,
    sheet: str,
    cell: str,
) -> tuple[Detector | None, TraupixeValidationIssue | None]:
    if raw_detector is None or (
        isinstance(raw_detector, str) and not raw_detector.strip()
    ):
        return None, None
    if not isinstance(raw_detector, str):
        return None, TraupixeValidationIssue(
            code=TraupixeValidationCode.UNKNOWN_DETECTOR,
            message="Detector labels must be text or empty",
            sheet=sheet,
            cell=cell,
        )
    return Detector(raw_detector.strip()), None


def _output_analysis_ids(
    analysis_order: tuple[AnalysisKey, ...],
) -> dict[AnalysisKey, str]:
    counts: dict[str, int] = {}
    for source_id, _ in analysis_order:
        counts[source_id] = counts.get(source_id, 0) + 1

    reserved_unique_ids = {
        source_id for source_id, count in counts.items() if count == 1
    }
    used: set[str] = set()
    output_ids: dict[AnalysisKey, str] = {}
    for analysis_key in analysis_order:
        source_id, occurrence = analysis_key
        if counts[source_id] == 1:
            candidate = source_id
        else:
            candidate = f"{source_id} [occurrence {occurrence}]"
            suffix = 1
            while candidate in reserved_unique_ids or candidate in used:
                suffix += 1
                candidate = f"{source_id} [occurrence {occurrence}; variant {suffix}]"
        used.add(candidate)
        output_ids[analysis_key] = candidate
    return output_ids


def _load_records(
    rows_by_sheet: dict[
        str,
        tuple[IdentifiedRow, ...],
    ],
    traupixe_format: TraupixeFormat,
    analytes: tuple[str, ...],
) -> tuple[
    tuple[Analysis, ...],
    tuple[LoadedMeasurement, ...],
    tuple[Detector, ...],
    list[TraupixeValidationIssue],
]:
    analysis_order = tuple(
        analysis_key
        for _, analysis_key, _ in rows_by_sheet[traupixe_format.unit_sheets[0][0]]
    )
    output_analysis_ids = _output_analysis_ids(analysis_order)
    row_maps = {
        sheet_name: {
            analysis_key: (row_number, row) for row_number, analysis_key, row in rows
        }
        for sheet_name, rows in rows_by_sheet.items()
    }

    analyses = tuple(
        Analysis(
            analysis_id=output_analysis_ids[analysis_key],
            description=(
                ""
                if row_maps["Exp. data"][analysis_key][1][1] is None
                else str(row_maps["Exp. data"][analysis_key][1][1])
            ),
        )
        for analysis_key in analysis_order
    )

    detector_rows = row_maps["S_Best Det."]
    detector_map: dict[tuple[AnalysisKey, str], Detector | None] = {}
    issues: list[TraupixeValidationIssue] = []
    used_detectors: list[Detector] = []
    for analysis_key in analysis_order:
        row_number, row = detector_rows[analysis_key]
        for analyte_index, analyte in enumerate(
            analytes,
        ):
            column_index = analyte_index + 3
            detector, issue = _normalize_detector(
                row[analyte_index + 2],
                sheet="S_Best Det.",
                cell=f"{_column_letter(column_index)}{row_number}",
            )
            if issue is not None:
                issues.append(issue)
                continue
            detector_map[(analysis_key, analyte)] = detector
            if detector is not None and detector not in used_detectors:
                used_detectors.append(detector)

    if issues:
        return analyses, (), (), issues

    measurements: list[LoadedMeasurement] = []
    unit_rows = {
        unit: row_maps[sheet_name] for sheet_name, unit in traupixe_format.unit_sheets
    }
    for analysis_key in analysis_order:
        for analyte_index, analyte in enumerate(
            analytes,
        ):
            value_index = 2 + (analyte_index * 2)
            uncertainty_index = value_index + 1
            for unit in traupixe_format.units:
                _, row = unit_rows[unit][analysis_key]
                measurements.append(
                    LoadedMeasurement(
                        analysis_id=output_analysis_ids[analysis_key],
                        analyte=analyte,
                        unit=unit,
                        raw_value=row[value_index],
                        raw_uncertainty=row[uncertainty_index],
                        detector=detector_map[(analysis_key, analyte)],
                    )
                )

    return analyses, tuple(measurements), tuple(used_detectors), []


def _column_letter(column: int) -> str:
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def load_traupixe_workbook(
    source: WorkbookSource,
    *,
    source_name: str | None = None,
    traupixe_format: TraupixeFormat = TRAUPIXE_FORMAT,
) -> LoadedTraupixeWorkbook:
    try:
        source_details = _read_source_details(
            source,
            source_name,
            traupixe_format.maximum_source_size,
        )
    except OSError as error:
        raise TraupixeUnreadableError(
            f"Unable to read TRAUPIXE source: {error}"
        ) from error
    extension = Path(source_details.name).suffix
    if extension.lower() != traupixe_format.extension:
        _restore_stream_position(source, source_details.initial_position)
        raise TraupixeUnsupportedFileError(extension)
    if source_details.size > traupixe_format.maximum_source_size:
        _restore_stream_position(source, source_details.initial_position)
        raise TraupixeTooLargeError(
            source_details.size,
            traupixe_format.maximum_source_size,
        )

    formula_workbook = None
    workbook = None
    try:
        formula_workbook = load_workbook(
            source,
            read_only=True,
            data_only=False,
        )
        issues = _validate_sheet_names(
            tuple(formula_workbook.sheetnames),
            traupixe_format,
        )
        if not issues:
            issues.extend(
                _validate_no_formulas(
                    formula_workbook,
                    traupixe_format,
                )
            )
        if issues:
            raise TraupixeIncompatibleWorkbookError(tuple(issues))
        formula_workbook.close()
        formula_workbook = None
        if not isinstance(source, (str, Path)):
            source.seek(0)

        workbook = load_workbook(
            source,
            read_only=True,
            data_only=True,
        )
        issues = _validate_sheet_names(
            tuple(workbook.sheetnames),
            traupixe_format,
        )
        if issues:
            raise TraupixeIncompatibleWorkbookError(tuple(issues))

        for worksheet_format in traupixe_format.worksheets:
            _prepare_worksheet(workbook[worksheet_format.name])
        analytes, header_issues = _discover_analytes(
            workbook,
            traupixe_format,
        )
        issues.extend(header_issues)
        if issues:
            raise TraupixeIncompatibleWorkbookError(tuple(issues))

        rows_by_sheet: dict[
            str,
            tuple[IdentifiedRow, ...],
        ] = {}
        for sheet_name in traupixe_format.source_sheets:
            worksheet_format = traupixe_format.worksheet(sheet_name)
            if sheet_name == "Exp. data":
                maximum_column = 2
            elif sheet_name == "S_Best Det.":
                maximum_column = 2 + len(analytes)
            else:
                maximum_column = 2 + (len(analytes) * 2)
            rows, row_issues = _identified_rows(
                workbook[sheet_name],
                worksheet_format,
                maximum_column=maximum_column,
            )
            rows_by_sheet[sheet_name] = rows
            issues.extend(row_issues)
        issues.extend(
            _validate_aligned_ids(
                rows_by_sheet,
                traupixe_format.source_sheets,
            )
        )
        issues.extend(
            _validate_measurement_value_types(
                rows_by_sheet,
                traupixe_format,
            )
        )
        if issues:
            raise TraupixeIncompatibleWorkbookError(tuple(issues))

        analyses, measurements, detectors, record_issues = _load_records(
            rows_by_sheet,
            traupixe_format,
            analytes,
        )
        if record_issues:
            raise TraupixeIncompatibleWorkbookError(tuple(record_issues))
        return LoadedTraupixeWorkbook(
            source_name=source_details.name,
            source_sha256=source_details.sha256,
            analyses=analyses,
            measurements=measurements,
            analytes=analytes,
            units=traupixe_format.units,
            detectors=detectors,
        )
    except TraupixeIncompatibleWorkbookError:
        raise
    except (
        BadZipFile,
        InvalidFileException,
        OSError,
        ValueError,
        KeyError,
    ) as error:
        raise TraupixeUnreadableError(
            f"Unable to read TRAUPIXE workbook: {error}"
        ) from error
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if workbook is not None:
            workbook.close()
        _restore_stream_position(source, source_details.initial_position)


def validate_traupixe_workbook(
    source: WorkbookSource,
    *,
    source_name: str | None = None,
    traupixe_format: TraupixeFormat = TRAUPIXE_FORMAT,
) -> None:
    load_traupixe_workbook(
        source,
        source_name=source_name,
        traupixe_format=traupixe_format,
    )
