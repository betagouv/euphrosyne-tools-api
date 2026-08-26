from __future__ import annotations

import math
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Literal
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

CONCENTRATION_SHEETS = ("S_Conc. %", "S_Conc. ppm")
BEST_DETECTOR_SHEET = "S_Best Det."
HEADER_SCAN_ROWS = 10
MIN_ANALYTE_COLUMNS = 3
POINT_SUFFIX = re.compile(r"(?:[_\s-]+(?:pt|point)\s*\d+)\s*$", re.IGNORECASE)
NON_DETECTION = re.compile(r"^\s*<\s*")
DECIMAL_VALUE = re.compile(r"[-+]?(?:\d+(?:[.,]\d*)?|[.,]\d+)(?:[Ee][-+]?\d+)?")

AnalysisKind = Literal["reference", "unknown"]
DetectorKey = tuple[str, int, str]


class TraupixeDatasetError(ValueError):
    """Raised when a workbook cannot be interpreted as a supported TRAUPIXE file."""


@dataclass(frozen=True)
class TraupixeMeasurement:
    value: float
    detected: bool
    detector: str | None


@dataclass(frozen=True)
class TraupixeAnalysis:
    identifier: str
    label: str
    group: str
    kind: AnalysisKind
    measurements: dict[str, TraupixeMeasurement]


@dataclass(frozen=True)
class TraupixeDataset:
    concentration_sheet: str
    detector_sheet: str | None
    unit: str
    analytes: tuple[str, ...]
    analyses: tuple[TraupixeAnalysis, ...]

    @property
    def non_reference_analyses(self) -> tuple[TraupixeAnalysis, ...]:
        return tuple(
            analysis for analysis in self.analyses if analysis.kind != "reference"
        )

    def descriptor(self) -> dict[str, Any]:
        selected_analyses = self.non_reference_analyses
        groups: dict[str, int] = {}
        for analysis in selected_analyses:
            groups[analysis.group] = groups.get(analysis.group, 0) + 1
        analytes = []
        for analyte in self.analytes:
            detected = sum(
                analysis.measurements[analyte].detected
                for analysis in selected_analyses
                if analyte in analysis.measurements
            )
            detectors = sorted(
                {
                    measurement.detector
                    for analysis in selected_analyses
                    if (measurement := analysis.measurements.get(analyte)) is not None
                    and measurement.detector is not None
                }
            )
            analytes.append(
                {
                    "name": analyte,
                    "detected": detected,
                    "total_non_reference_analyses": len(selected_analyses),
                    "detection_rate": (
                        round(detected / len(selected_analyses), 4)
                        if selected_analyses
                        else 0
                    ),
                    "detectors": detectors,
                }
            )
        return {
            "concentration_sheet": self.concentration_sheet,
            "detector_sheet": self.detector_sheet,
            "unit": self.unit,
            "analyses": {
                "total": len(self.analyses),
                "references": sum(
                    analysis.kind == "reference" for analysis in self.analyses
                ),
                "unknown": sum(
                    analysis.kind == "unknown" for analysis in self.analyses
                ),
                "groups": [
                    {"name": name, "analyses": count} for name, count in groups.items()
                ],
            },
            "analytes": analytes,
        }


def serialize_traupixe_for_model(dataset: TraupixeDataset) -> dict[str, Any]:
    """Serialize normalized TRAUPIXE data into the compact model contract."""
    descriptor = dataset.descriptor()
    analyses = []
    for analysis in dataset.analyses:
        measurements = [
            analysis.measurements.get(analyte) for analyte in dataset.analytes
        ]
        analyses.append(
            {
                "identifier": analysis.identifier,
                "label": analysis.label,
                "group": analysis.group,
                "kind": analysis.kind,
                "values": [
                    measurement.value if measurement is not None else None
                    for measurement in measurements
                ],
                "detected": [
                    measurement.detected if measurement is not None else None
                    for measurement in measurements
                ],
                "detectors": [
                    measurement.detector if measurement is not None else None
                    for measurement in measurements
                ],
            }
        )
    return {
        "concentration_sheet": dataset.concentration_sheet,
        "detector_sheet": dataset.detector_sheet,
        "unit": dataset.unit,
        "summary": descriptor["analyses"],
        "analytes": descriptor["analytes"],
        "analyses": analyses,
    }


def load_traupixe_dataset(content: bytes) -> TraupixeDataset:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise TraupixeDatasetError("Unable to read the TRAUPIXE workbook") from error

    try:
        concentration_sheet = _find_sheet(workbook.sheetnames, CONCENTRATION_SHEETS)
        if concentration_sheet is None:
            raise TraupixeDatasetError(
                "No supported TRAUPIXE concentration sheet was found"
            )
        concentration = workbook[concentration_sheet]
        header_row = _find_concentration_header_row(concentration)
        analyte_columns = _analyte_columns(concentration, header_row)
        detector_sheet = _find_sheet(workbook.sheetnames, (BEST_DETECTOR_SHEET,))
        detector_values = (
            _detector_values(workbook[detector_sheet], tuple(analyte_columns))
            if detector_sheet is not None
            else {}
        )
        analyses = _analyses(
            concentration,
            header_row,
            analyte_columns,
            detector_values,
        )
    finally:
        workbook.close()

    if not analyses:
        raise TraupixeDatasetError("The TRAUPIXE workbook contains no analyses")
    return TraupixeDataset(
        concentration_sheet=concentration_sheet,
        detector_sheet=detector_sheet,
        unit="%" if concentration_sheet.casefold().endswith("%") else "ppm",
        analytes=tuple(analyte_columns),
        analyses=tuple(analyses),
    )


def _find_sheet(sheet_names: list[str], candidates: tuple[str, ...]) -> str | None:
    by_name = {_normalized_text(name): name for name in sheet_names}
    return next(
        (
            by_name[_normalized_text(candidate)]
            for candidate in candidates
            if _normalized_text(candidate) in by_name
        ),
        None,
    )


def _find_concentration_header_row(sheet: Worksheet) -> int:
    candidates = []
    for row in range(1, min(sheet.max_row, HEADER_SCAN_ROWS) + 1):
        if not _is_blank(sheet.cell(row, 1).value) or not _is_blank(
            sheet.cell(row, 2).value
        ):
            continue
        populated = sum(
            isinstance(sheet.cell(row, column).value, str)
            and bool(str(sheet.cell(row, column).value).strip())
            for column in range(3, sheet.max_column + 1)
        )
        candidates.append((populated, row))
    populated, row = max(candidates, key=lambda candidate: candidate[0], default=(0, 0))
    if populated < MIN_ANALYTE_COLUMNS:
        raise TraupixeDatasetError("Unable to locate the TRAUPIXE analyte header")
    return row


def _find_detector_header_row(
    sheet: Worksheet,
    analytes: tuple[str, ...],
) -> int | None:
    expected = set(analytes)
    candidates = []
    for row in range(1, min(sheet.max_row, HEADER_SCAN_ROWS) + 1):
        if not _is_blank(sheet.cell(row, 1).value) or not _is_blank(
            sheet.cell(row, 2).value
        ):
            continue
        overlap = sum(
            _text(sheet.cell(row, column).value) in expected
            for column in range(3, sheet.max_column + 1)
        )
        candidates.append((overlap, row))
    overlap, row = max(candidates, key=lambda candidate: candidate[0], default=(0, 0))
    return row if overlap else None


def _analyte_columns(sheet: Worksheet, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column in range(3, sheet.max_column + 1):
        value = sheet.cell(header_row, column).value
        if isinstance(value, str) and value.strip():
            columns.setdefault(value.strip(), column)
    if len(columns) < MIN_ANALYTE_COLUMNS:
        raise TraupixeDatasetError("The TRAUPIXE analyte header is incomplete")
    return columns


def _detector_values(
    sheet: Worksheet,
    analytes: tuple[str, ...],
) -> dict[DetectorKey, str]:
    header_row = _find_detector_header_row(sheet, analytes)
    if header_row is None:
        return {}
    headers = {
        header: column
        for column in range(3, sheet.max_column + 1)
        if (header := _text(sheet.cell(header_row, column).value)) is not None
    }
    matching_columns = {
        analyte: headers[analyte] for analyte in analytes if analyte in headers
    }
    values: dict[DetectorKey, str] = {}
    occurrences: dict[str, int] = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        identifier = _text(sheet.cell(row, 1).value)
        if identifier is None:
            continue
        occurrence = occurrences.get(identifier, 0)
        occurrences[identifier] = occurrence + 1
        for analyte, column in matching_columns.items():
            detector = _text(sheet.cell(row, column).value)
            if detector is not None:
                values[(identifier, occurrence, analyte)] = detector
    return values


def _analyses(
    sheet: Worksheet,
    header_row: int,
    analyte_columns: dict[str, int],
    detector_values: dict[DetectorKey, str],
) -> list[TraupixeAnalysis]:
    analyses = []
    occurrences: dict[str, int] = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        identifier = _text(sheet.cell(row, 1).value)
        if identifier is None:
            continue
        occurrence = occurrences.get(identifier, 0)
        occurrences[identifier] = occurrence + 1
        label = _text(sheet.cell(row, 2).value)
        if label is None:
            continue
        measurements = {}
        for analyte, column in analyte_columns.items():
            parsed = _parse_measurement(sheet.cell(row, column).value)
            if parsed is None:
                continue
            value, detected = parsed
            measurements[analyte] = TraupixeMeasurement(
                value=value,
                detected=detected,
                detector=detector_values.get((identifier, occurrence, analyte)),
            )
        if not measurements:
            continue
        analyses.append(
            TraupixeAnalysis(
                identifier=identifier,
                label=label,
                group=_analysis_group(label),
                kind=_analysis_kind(identifier),
                measurements=measurements,
            )
        )
    return analyses


def _parse_measurement(value: Any) -> tuple[float, bool] | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (float(value), True) if math.isfinite(value) else None
    if not isinstance(value, str):
        return None
    match = DECIMAL_VALUE.search(value.strip())
    if match is None:
        return None
    parsed = float(match.group(0).replace(",", "."))
    if not math.isfinite(parsed):
        return None
    return parsed, NON_DETECTION.match(value) is None


def _analysis_group(label: str) -> str:
    group = POINT_SUFFIX.sub("", label).rstrip(" _-")
    return group or label


def _analysis_kind(identifier: str) -> AnalysisKind:
    return "reference" if "_std_" in identifier.casefold() else "unknown"


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    stripped = value.strip()
    return stripped or None


def _normalized_text(value: str) -> str:
    return " ".join(value.strip().casefold().split())
